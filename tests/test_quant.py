"""
정량 회귀 테스트 (통합, 가드레일)
---------------------------------
기존 시트의 56개 조성을 후보로 고정하고, 도구의 상대정량(%) 순위가
시트의 상대량(%) 순위와 Spearman >= 0.55 로 상관하는지 검증.
(절대값 일치는 기대하지 않음 — 사람이 읽은 NL/면적값과 방식이 달라 순위로만 평가)

샘플 mzML(tests/data/sample.mzML)이 없으면 skip.
실행:  pytest tests/test_quant.py
"""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pytest

from glyco import masses, compositions, mzml_parse, identify

HERE = os.path.dirname(os.path.abspath(__file__))
MZML = os.path.join(HERE, "data", "sample.mzML")
FIXTURE = os.path.join(HERE, "data", "calculator_fixture.json")
SHEET = "/Users/hitodi/Documents/x-callibur/Alzhemer of brain.xlsx"
MIN_SPEARMAN = 0.57   # area 정량 기준(현재 +0.60); 회귀 방지 하한


def _spearman(a, b):
    n = len(a)
    def rank(x):
        order = sorted(range(n), key=lambda i: x[i])
        rk = [0] * n
        for p, i in enumerate(order):
            rk[i] = p
        return rk
    ra, rb = rank(a), rank(b)
    d2 = sum((ra[i] - rb[i]) ** 2 for i in range(n))
    return 1 - 6 * d2 / (n * (n * n - 1))


def _sheet_pct():
    import openpyxl
    ws = openpyxl.load_workbook(SHEET, data_only=True)["Relative Quantity"]
    pct = {}
    r = 9
    while r <= 189:
        no, ak = ws.cell(r, 4).value, ws.cell(r, 37).value
        if isinstance(no, (int, float)) and isinstance(ak, (int, float)):
            pct[int(no)] = float(ak)
        r += 4
    return pct


@pytest.mark.skipif(not os.path.exists(MZML), reason="sample.mzML 없음(대용량, 미커밋)")
@pytest.mark.skipif(not os.path.exists(SHEET), reason="원본 엑셀 없음")
def test_quant_rank_correlation():
    fix = json.load(open(FIXTURE))
    keyc = lambda c: (c["HexNAc"], c["Hex"], c["dHex"], c["Neu5Ac"], c["Neu5Gc"], c.get("Xyl", 0))
    cands, no_by_key = [], {}
    for rec in fix:
        c = {k: rec[k] for k in ("HexNAc", "Hex", "dHex", "Neu5Ac", "Neu5Gc", "Xyl", "ProA")}
        cands.append({"composition": c, "neutral": masses.neutral_mass(c),
                      "formula": masses.formula_str(c), "name": compositions.composition_name(c)})
        no_by_key[keyc(c)] = rec["no"]

    data = mzml_parse.parse(MZML, log=lambda *a: None)
    res = identify.run(cands, msdata=data, ppm_tol=10, ms1_ppm=5,
                       require_ms2=True, log=lambda *a: None)

    sheet = _sheet_pct()
    rows = [(no_by_key.get(keyc(r["composition"])), r["relative_pct"]) for r in res]
    rows = [(no, p) for no, p in rows if no in sheet]
    rho = _spearman([sheet[no] for no, _ in rows], [p for _, p in rows])
    print(f"[정량] 매칭 {len(rows)}개, Spearman = {rho:+.3f}")
    assert rho >= MIN_SPEARMAN, f"Spearman {rho:.3f} < {MIN_SPEARMAN} (정량 회귀 의심)"


if __name__ == "__main__":
    test_quant_rank_correlation()

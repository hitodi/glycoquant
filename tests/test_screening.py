import os
import sys

import numpy as np
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from glyco import identify, report
from glyco.mzml_parse import MsData


def test_screening_table_filters_anchor_and_keeps_precursor_info(tmp_path):
    data = MsData()
    data.ms2 = [
        (10.0, 527.2279, 2, "1001"),
        (11.0, 600.0000, 1, "1002"),
    ]
    data.ms2_peaks = {
        "1001": (
            np.array([204.0867, 366.1395, 441.2708]),
            np.array([1000.0, 200.0, 800.0]),
        ),
        "1002": (
            np.array([366.1395, 441.2708]),
            np.array([200.0, 800.0]),
        ),
    }
    data.ms2_info = {"1001": {"isolation_target_mz": 527.2277}}

    rows, ions = identify.screening_table(
        data,
        [("HexNAc", 204.0867), ("HexNAc+Hex", 366.1395), ("ProA-HexNAc", 441.2708)],
        ppm=20.0,
        anchor="HexNAc",
    )

    assert ions == ["HexNAc", "HexNAc+Hex", "ProA-HexNAc"]
    assert len(rows) == 1
    assert rows[0]["scan"] == "1001"
    assert rows[0]["charge"] == 2
    assert rows[0]["isolation_target"] == 527.2277
    assert rows[0]["monoisotope"] == 527.2279
    assert rows[0]["ions"]["HexNAc"] == 204.0867

    out = tmp_path / "screening.xlsx"
    report.write_screening(rows, ions, out, meta={"sample": "sample.raw", "source": "test"})
    wb = load_workbook(out, data_only=True)
    ws = wb["Screening"]
    assert ws["A1"].value.startswith("Per-MS2-scan")
    assert ws["A3"].value == "Scan No"
    assert ws["A4"].value == 1001
    assert ws["F4"].value == 527.2277
    assert ws["G4"].value == 527.2279


def test_screening_table_accepts_comma_separated_anchor_list():
    data = MsData()
    data.ms2 = [(10.0, 527.2279, 2, "1001")]
    data.ms2_peaks = {
        "1001": (
            np.array([441.2708]),
            np.array([800.0]),
        ),
    }

    rows, _ = identify.screening_table(
        data,
        [("HexNAc", 204.0867), ("ProA-HexNAc", 441.2708)],
        ppm=20.0,
        anchor="HexNAc,ProA-HexNAc",
    )

    assert len(rows) == 1
    assert rows[0]["ions"]["HexNAc"] is None
    assert rows[0]["ions"]["ProA-HexNAc"] == 441.2708

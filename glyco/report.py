"""
엑셀 리포트 작성
----------------
시트 3개:
  - Glycans  : 동정/정량된 글리칸 전체 표 (상대량 내림차순)
  - Summary  : 구조 유형별(High-mannose/Hybrid/Complex) + 시알/푸코 집계
  - Adducts  : 글리칸별 adduct 상세(검출된 이온 종류·m/z·강도)
상대량(%)·집계는 하드코딩이 아니라 엑셀 수식으로 넣어 재계산 가능하게 한다.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT, bold=True, size=13)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_screening_sheet(wb, screening, screening_ions, *, title="Screening", meta=None):
    """Xcalibur 수작업 스크리닝 전사용 시트를 추가한다."""
    sc = wb.create_sheet(title)
    sc.append(["Per-MS2-scan diagnostic ions (Xcalibur 수작업 대체)"])
    sc["A1"].font = TITLE_FONT
    if meta:
        sc.append([f"Sample: {meta.get('sample','')}   |   Source: {meta.get('source','')}"])
        sc["A2"].font = Font(name=FONT, italic=True, color="666666")
    hdr_row = sc.max_row + 1
    hdr = ["Scan No", "RT (min)"] + [f"{n}\n(obs m/z)" for n in screening_ions] \
        + ["Precursor m/z", "Monoisotope/selected m/z", "Charge", "비고"]
    sc.append(hdr)
    _style_header(sc, hdr_row, len(hdr))
    for row in screening or []:
        line = [int(row["scan"]) if str(row["scan"]).isdigit() else row["scan"],
                round(row["rt"], 2)]
        for n in screening_ions:
            v = row["ions"].get(n)
            line.append(round(v, 4) if v else None)
        precursor = row.get("isolation_target") or row["precursor"]
        mono = row.get("monoisotope")
        line += [
            round(precursor, 4),
            round(mono, 4) if isinstance(mono, (int, float)) else None,
            row["charge"] if row["charge"] else "?",
            None,
        ]
        sc.append(line)
    for rr in range(hdr_row + 1, sc.max_row + 1):
        for cc in range(1, len(hdr) + 1):
            sc.cell(rr, cc).font = Font(name=FONT)
            sc.cell(rr, cc).border = BORDER
    _autofit(sc, [9, 8] + [13] * len(screening_ions) + [14, 19, 7, 14])
    sc.freeze_panes = sc.cell(hdr_row + 1, 1)
    return sc


def write_screening(screening, screening_ions, out_path, meta=None):
    """Screening 전용 workbook 을 저장한다."""
    wb = Workbook()
    wb.remove(wb.active)
    _add_screening_sheet(wb, screening, screening_ions, meta=meta)
    wb.save(out_path)
    return out_path


def write_targeted(rows, spec, out_path, meta=None):
    """
    타깃 스크리닝 결과 workbook.
      - Matched (≥K) : 채택된 (scan,glycan) 쌍
      - Holding (1..K-1) : 보류(부분매칭)
      - Precursor groups : Matched 스캔을 precursor floor(step) 로 묶어 나열
    """
    from .identify import floor_bin
    step = spec.precursor_floor
    matched = [r for r in rows if r["tier"] == "matched"]
    holding = [r for r in rows if r["tier"] == "holding"]
    wb = Workbook()
    wb.remove(wb.active)

    # ---------- Glycan summary (0 포함 — 손입력 오류/누락 구분용) ----------
    from collections import Counter
    mc = Counter(r["glycan"] for r in matched)
    hc = Counter(r["glycan"] for r in holding)
    gs = wb.create_sheet("Glycan summary")
    gs.append([f"글리칸별 매칭 요약 (K={spec.min_hits}, ppm={spec.ppm})"])
    gs["A1"].font = TITLE_FONT
    gs.append(["Glycan", "이온 N", "Matched", "Holding", "비고"])
    _style_header(gs, 2, 5)
    for g in spec.glycans:
        m, h = mc.get(g["name"], 0), hc.get(g["name"], 0)
        note = "⚠ 매칭 0 — 입력 m/z 확인" if m == 0 else ""
        gs.append([g["name"], len(g["ions"]), m, h, note])
    for rr in range(3, gs.max_row + 1):
        for cc in range(1, 6):
            gs.cell(rr, cc).font = Font(name=FONT); gs.cell(rr, cc).border = BORDER
    _autofit(gs, [18, 8, 10, 10, 24])

    def _sheet(title, data):
        ws = wb.create_sheet(title)
        ws.append([title])
        ws["A1"].font = TITLE_FONT
        hdr = ["Glycan", "Scan No", "RT (min)", "Hits (n/N)", "Precursor m/z", "Charge", "Matched ions (obs m/z)"]
        ws.append(hdr)
        _style_header(ws, 2, len(hdr))
        for r in data:
            hit_ions = "; ".join(f"{t:.4f}→{o:.4f}" for t, o in r["ion_obs"].items() if o is not None)
            ws.append([r["glycan"],
                       int(r["scan"]) if str(r["scan"]).isdigit() else r["scan"],
                       round(r["rt"], 2), f"{r['n_hit']}/{r['n_ions']}",
                       round(r["precursor"], 4), r["charge"] if r["charge"] else "?", hit_ions])
        for rr in range(3, ws.max_row + 1):
            for cc in range(1, len(hdr) + 1):
                ws.cell(rr, cc).font = Font(name=FONT); ws.cell(rr, cc).border = BORDER
        _autofit(ws, [16, 9, 9, 10, 14, 7, 46])
        ws.freeze_panes = "A3"
        return ws

    _sheet(f"Matched (≥{spec.min_hits})", matched)
    _sheet("Holding (부분매칭)", holding)

    # Precursor groups — Matched 스캔을 floor(step) 로 묶음(표시는 전체값)
    pg = wb.create_sheet("Precursor groups")
    pg.append([f"Matched precursor 그룹 (floor {step} 단위)"])
    pg["A1"].font = TITLE_FONT
    hdr = ["Group (floor)", "Precursor m/z", "Glycan", "Scan No", "RT (min)", "Charge"]
    pg.append(hdr)
    _style_header(pg, 2, len(hdr))
    from collections import defaultdict
    groups = defaultdict(list)
    for r in matched:
        groups[floor_bin(r["precursor"], step)].append(r)
    for key in sorted(groups):
        for r in sorted(groups[key], key=lambda x: x["precursor"]):
            pg.append([round(key, 4), round(r["precursor"], 4), r["glycan"],
                       int(r["scan"]) if str(r["scan"]).isdigit() else r["scan"],
                       round(r["rt"], 2), r["charge"] if r["charge"] else "?"])
    for rr in range(3, pg.max_row + 1):
        for cc in range(1, len(hdr) + 1):
            pg.cell(rr, cc).font = Font(name=FONT); pg.cell(rr, cc).border = BORDER
    _autofit(pg, [14, 14, 16, 9, 9, 7])
    pg.freeze_panes = "A3"

    wb.save(out_path)
    return out_path


def write_aggregated(agg, out_path, meta=None):
    """
    반복 취합 결과를 workbook 으로 저장한다.
      - Aggregated : 글리칸 × [평균/SD/CV/검출빈도 + 반복별 %]
      - Type summary : 유형별 평균±SD (반복별 %)
    """
    files = agg["files"]
    wb = Workbook()

    # ---------- Aggregated ----------
    ws = wb.active
    ws.title = "Aggregated"
    ws.append([f"반복 취합 결과 (n={len(files)})"])
    ws["A1"].font = TITLE_FONT
    if meta:
        ws.append([f"그룹: {meta.get('group', '')}   |   파일: {', '.join(files)}"])
        ws["A2"].font = Font(name=FONT, italic=True, color="666666")
    hr = ws.max_row + 1
    base = ["No.", "Oxford", "Composition", "Type", "Evidence",
            "Mean %", "SD", "CV %", "n/N (검출)"]
    headers = base + [f"{f}\n%" for f in files]
    ws.append(headers)
    _style_header(ws, hr, len(headers))

    from .compositions import composition_name
    _r = lambda v, n: (round(v, n) if v is not None else "-")
    for i, g in enumerate(agg["glycans"]):
        row = [
            i + 1, g["oxford"], composition_name(g["composition"]), g["type"], g["evidence"],
            round(g["mean"], 2), _r(g["sd"], 2), _r(g["cv"], 1),
            f"{g['n_detected']}/{g['n_total']}",
        ]
        for f in files:
            v = g["pct_by_file"].get(f)
            row.append(round(v, 2) if v is not None else None)
        ws.append(row)
    last = ws.max_row
    for rr in range(hr + 1, last + 1):
        for cc in range(1, len(headers) + 1):
            ws.cell(rr, cc).font = Font(name=FONT)
            ws.cell(rr, cc).border = BORDER
    _autofit(ws, [5, 12, 26, 13, 11, 9, 8, 8, 11] + [11] * len(files))
    ws.freeze_panes = ws.cell(hr + 1, 1)

    # ---------- Type summary ----------
    ts = wb.create_sheet("Type summary")
    ts.append(["유형별 취합 (평균 ± SD, %)"])
    ts["A1"].font = TITLE_FONT
    th = ["Type", "Mean %", "SD"] + [f"{f} %" for f in files]
    ts.append(th)
    _style_header(ts, 2, len(th))
    for t in ("High-mannose", "Hybrid", "Complex"):
        d = agg["types"][t]
        sd = round(d["sd"], 1) if d["sd"] is not None else "-"
        row = [t, round(d["mean"], 1), sd] + [round(d["by_file"][f], 1) for f in files]
        ts.append(row)
    for rr in range(3, ts.max_row + 1):
        for cc in range(1, len(th) + 1):
            ts.cell(rr, cc).font = Font(name=FONT)
            ts.cell(rr, cc).border = BORDER
    _autofit(ts, [14, 9, 8] + [11] * len(files))

    wb.save(out_path)
    return out_path


def write(results, out_path, meta=None, screening=None, screening_ions=None):
    wb = Workbook()

    # ---------- Glycans ----------
    ws = wb.active
    ws.title = "Glycans"
    headers = ["No.", "Oxford", "Name (composition)", "HexNAc", "Hex", "dHex/Fuc",
               "Neu5Ac", "Neu5Gc", "Type", "Sialylated", "Fucosylated",
               "Best ion m/z", "Adduct", "Charge", "RT (min)",
               "MS2 #", "Evidence", "ppm", "Quant (sum)", "Relative %"]
    ws.append(["N-glycan analysis result"])
    ws["A1"].font = TITLE_FONT
    if meta:
        ws.append([f"Sample: {meta.get('sample','')}   |   Source: {meta.get('source','')}"])
        ws["A2"].font = Font(name=FONT, italic=True, color="666666")
    hr = ws.max_row + 1
    ws.append(headers)
    _style_header(ws, hr, len(headers))

    first = hr + 1
    # 헤더명으로 열 위치를 동적 계산(헤더 변경에 견고)
    col = lambda name: headers.index(name) + 1
    let = lambda name: get_column_letter(col(name))
    inten_col = col("Quant (sum)")
    inten_letter = let("Quant (sum)")
    for i, r in enumerate(results):
        c = r["composition"]
        row = [
            i + 1, r.get("oxford", ""), r["name"], c["HexNAc"], c["Hex"], c["dHex"],
            c["Neu5Ac"], c["Neu5Gc"], r["type"],
            "Y" if r["sialylated"] else "", "Y" if r["fucosylated"] else "",
            round(r["best_mz"], 4) if r["best_mz"] else None,
            r["best_adduct"], r["best_z"],
            round(r["rt"], 2) if r["rt"] is not None else None,
            r["ms2_count"],
            r.get("evidence", "MS2"),
            round(r["best_ppm"], 2) if r["best_ppm"] is not None else None,
            r["intensity_sum"], None,
        ]
        ws.append(row)
    last = ws.max_row
    total_row = last + 1
    # Relative % = intensity / total * 100  (수식)
    if results:
        for rr in range(first, last + 1):
            cell = ws.cell(rr, len(headers))
            cell.value = f"={inten_letter}{rr}/{inten_letter}${total_row}*100"
            cell.number_format = "0.00"
        ws.cell(total_row, 1, "Total")
        ws.cell(total_row, 1).font = Font(name=FONT, bold=True)
        ws.cell(total_row, inten_col,
                f"=SUM({inten_letter}{first}:{inten_letter}{last})").font = Font(name=FONT, bold=True)
        ws.cell(total_row, len(headers),
                f"=SUM({get_column_letter(len(headers))}{first}:{get_column_letter(len(headers))}{last})")
        ws.cell(total_row, len(headers)).number_format = "0.00"

    for rr in range(first, last + 1):
        for cc in range(1, len(headers) + 1):
            cell = ws.cell(rr, cc)
            cell.font = Font(name=FONT)
            cell.border = BORDER
            if cc in (inten_col,):
                cell.number_format = "#,##0"
    _autofit(ws, [5, 12, 28, 8, 7, 9, 8, 8, 13, 10, 11, 13, 10, 8, 9, 7, 10, 7, 16, 11])
    ws.freeze_panes = ws.cell(first, 1)

    # ---------- Summary ----------
    ss = wb.create_sheet("Summary")
    ss.append(["Structure-type summary"])
    ss["A1"].font = TITLE_FONT
    ss.append(["Type", "Count", "Relative % (sum)"])
    _style_header(ss, 2, 3)
    types = ["High-mannose", "Hybrid", "Complex"]
    gl = "Glycans"
    tl, pl = let("Type"), let("Relative %")          # 동적 열 문자
    type_range = f"'{gl}'!{tl}{first}:{tl}{last}"
    pct_range = f"'{gl}'!{pl}{first}:{pl}{last}"
    for t in types:
        ss.append([t, None, None])
        rr = ss.max_row
        ss.cell(rr, 2, f'=COUNTIF({type_range},"{t}")')
        ss.cell(rr, 3, f'=SUMIF({type_range},"{t}",{pct_range})')
        ss.cell(rr, 3).number_format = "0.00"
    base = ss.max_row - 2
    ss.append(["Sum", f"=SUM(B{base}:B{base+2})", f"=SUM(C{base}:C{base+2})"])
    ss.cell(ss.max_row, 3).number_format = "0.00"
    for rr in range(ss.max_row, ss.max_row + 1):
        for cc in range(1, 4):
            ss.cell(rr, cc).font = Font(name=FONT, bold=True)

    ss.append([])
    ss.append(["Modification", "Count", "Relative % (sum)"])
    _style_header(ss, ss.max_row, 3)
    sl, fl = let("Sialylated"), let("Fucosylated")
    sia_range = f"'{gl}'!{sl}{first}:{sl}{last}"
    fuc_range = f"'{gl}'!{fl}{first}:{fl}{last}"
    for label, rng in (("Sialylated", sia_range), ("Fucosylated", fuc_range)):
        ss.append([label, None, None])
        rr = ss.max_row
        ss.cell(rr, 2, f'=COUNTIF({rng},"Y")')
        ss.cell(rr, 3, f'=SUMIF({rng},"Y",{pct_range})')
        ss.cell(rr, 3).number_format = "0.00"
    for rr in range(2, ss.max_row + 1):
        for cc in range(1, 4):
            cell = ss.cell(rr, cc)
            if cell.value is not None and not (rr == 2):
                cell.border = BORDER
                if cell.font.name != FONT:
                    cell.font = Font(name=FONT)
    _autofit(ss, [16, 10, 16])

    # ---------- Adducts ----------
    ad = wb.create_sheet("Adducts")
    ad.append(["Per-glycan detected adduct ions"])
    ad["A1"].font = TITLE_FONT
    ad.append(["Glycan", "Type", "Adduct", "Charge", "Theo m/z", "RT (min)", "Intensity"])
    _style_header(ad, 2, 7)
    for r in results:
        for a in sorted(r["adducts"], key=lambda x: -x["intensity"]):
            ad.append([r["name"], r["type"], a["adduct"], a["z"],
                       round(a["mz"], 4), round(a["rt"], 2) if a["rt"] == a["rt"] else None,
                       a["intensity"]])
    for rr in range(3, ad.max_row + 1):
        for cc in range(1, 8):
            cell = ad.cell(rr, cc)
            cell.font = Font(name=FONT)
            cell.border = BORDER
            if cc == 7:
                cell.number_format = "#,##0"
    _autofit(ad, [30, 13, 10, 8, 13, 10, 16])
    ad.freeze_panes = "A3"

    # ---------- Screening (Xcalibur 수작업 대체: 스캔별 진단이온/precursor) ----------
    if screening is not None and screening_ions:
        _add_screening_sheet(wb, screening, screening_ions, meta=meta)

    wb.save(out_path)
    return out_path
